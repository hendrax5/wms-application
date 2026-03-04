import io
from typing import Optional

from fastapi import HTTPException, UploadFile
from sqlalchemy.orm import Session

from app.models.device_instance import DeviceInstance
from app.models.device_model import DeviceModel
from app.models.movement_log import MovementLog
from app.models.location import Location


def process_bulk_import(db: Session, file: UploadFile, location_id: int) -> dict:
    """
    Process a CSV or Excel file to bulk-create DeviceInstances.

    Expected columns: serial_number, model_id, mac_address (optional)
    All devices are created as AVAILABLE at the given location.
    """
    filename = file.filename or ""
    content = file.file.read()

    rows = []
    if filename.endswith(".csv"):
        rows = _parse_csv(content)
    elif filename.endswith((".xlsx", ".xls")):
        rows = _parse_excel(content)
    else:
        raise HTTPException(status_code=400, detail="File must be .csv or .xlsx")

    # Validate location
    loc = db.query(Location).filter(Location.id == location_id).first()
    if not loc:
        raise HTTPException(status_code=404, detail=f"Location #{location_id} not found")

    results = {"success": 0, "errors": []}

    for i, row in enumerate(rows, start=2):  # start=2 because row 1 is header
        sn = row.get("serial_number", "").strip()
        model_id_str = row.get("model_id", "").strip()
        mac = row.get("mac_address", "").strip()

        if not sn or not model_id_str:
            results["errors"].append({"row": i, "error": "Missing serial_number or model_id"})
            continue

        try:
            model_id = int(model_id_str)
        except ValueError:
            results["errors"].append({"row": i, "error": f"Invalid model_id: {model_id_str}"})
            continue

        # Check model exists
        if not db.query(DeviceModel).filter(DeviceModel.id == model_id).first():
            results["errors"].append({"row": i, "error": f"DeviceModel #{model_id} not found"})
            continue

        # Check duplicate SN
        if db.query(DeviceInstance).filter(DeviceInstance.serial_number == sn).first():
            results["errors"].append({"row": i, "error": f"Duplicate serial_number: {sn}"})
            continue

        device = DeviceInstance(
            model_id=model_id,
            serial_number=sn,
            mac_address=mac or None,
            current_location_id=location_id,
            status="AVAILABLE",
            condition="GOOD",
        )
        db.add(device)
        db.flush()

        log = MovementLog(
            device_instance_id=device.id,
            from_location_id=None,
            to_location_id=location_id,
            activity_type="INBOUND",
            notes="Bulk import",
        )
        db.add(log)
        results["success"] += 1

    db.commit()
    return results


def _parse_csv(content: bytes) -> list[dict]:
    """Parse CSV content into list of dicts."""
    text = content.decode("utf-8-sig")
    lines = text.strip().split("\n")
    if len(lines) < 2:
        return []
    headers = [h.strip().lower() for h in lines[0].split(",")]
    rows = []
    for line in lines[1:]:
        values = [v.strip() for v in line.split(",")]
        row = dict(zip(headers, values))
        rows.append(row)
    return rows


def _parse_excel(content: bytes) -> list[dict]:
    """Parse Excel content into list of dicts using openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip().lower() for h in next(rows_iter)]
    rows = []
    for row_values in rows_iter:
        row = dict(zip(headers, [str(v).strip() if v is not None else "" for v in row_values]))
        rows.append(row)
    return rows
